#!/usr/bin/env python3
"""
Insere no banco os AIOs da planilha "Informações AIO.xlsx" (aba ListaAIO) que
NÃO existem em se_cgpac.aio_solicitacoes — os que o pipeline do Gmail não trouxe.

Os AIOs que já existem (casam por instrumento/tc) são PULADOS aqui; para eles,
use importar_datas_planilha.py (que atualiza as 3 datas das etapas).

Campos preenchidos a partir da planilha:
    instrumento        <- D  "Instrumento" (texto original, truncado em 20)
    objeto             <- F  "Local da Obra"
    municipio / uf     <- E  "Município/UF" (parse best-effort)
    valor_investimento <- G  "Valor Repasse da Etapa de AIO"
    dt_checklist_cgpac <- I  "Data de chegada na CGPAC"  (etapa 6)
    dt_gabse_oficio    <- J  "Data de entrega para SE"   (etapa 7)
    dt_assinatura_sex  <- K  "Data Assinatura GM"        (etapa 8)
    email_id           <- "PLANILHA-<digitos>" (chave sintética, idempotente)

Demais campos (etapas 1-5/9-10, programa, objeto detalhado) ficam vazios.

Uso:
    python3 importar_aios_planilha.py --arquivo "/caminho/Informações AIO.xlsx"            # dry-run
    python3 importar_aios_planilha.py --arquivo "..." --commit                             # grava
"""

import argparse
import datetime
import pathlib
import re
import sys

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from aio_pipeline import get_db_connection  # noqa: E402

ABA_PADRAO = "ListaAIO"
LINHA_CABECALHO = 2
COLS = {"instr": "D", "sei": "C", "munuf": "E", "local": "F", "valor": "G",
        "I": "I", "J": "J", "K": "K"}
MAPA_DATAS = {"I": "dt_checklist_cgpac", "J": "dt_gabse_oficio", "K": "dt_assinatura_sex"}

UFS = {"AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG","PA",
       "PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"}

# Nomes de estado por extenso -> sigla (alguns registros vêm como "Estado/Cidade")
ESTADOS = {
    "acre":"AC","alagoas":"AL","amapa":"AP","amazonas":"AM","bahia":"BA",
    "ceara":"CE","distrito federal":"DF","espirito santo":"ES","goias":"GO",
    "maranhao":"MA","mato grosso":"MT","mato grosso do sul":"MS","minas gerais":"MG",
    "para":"PA","paraiba":"PB","parana":"PR","pernambuco":"PE","piaui":"PI",
    "rio de janeiro":"RJ","rio grande do norte":"RN","rio grande do sul":"RS",
    "rondonia":"RO","roraima":"RR","santa catarina":"SC","sao paulo":"SP",
    "sergipe":"SE","tocantins":"TO",
}


def _sem_acento(s):
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFKD", s)
                   if not unicodedata.combining(c))


def _col(l):
    return openpyxl.utils.column_index_from_string(l)


def _digitos(v):
    if v is None:
        return ""
    return re.sub(r"\D", "", str(v).split("/")[0])


def _as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def _parse_valor(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d,.-]", "", str(v))          # tira "R$", espaços, \xa0
    s = s.replace(".", "").replace(",", ".")     # 1.299.609,84 -> 1299609.84
    try:
        return float(s)
    except ValueError:
        return None


def _parse_munuf(v):
    """Best-effort: separa município e UF de 'Município/UF' (ou 'UF/Município')."""
    if not v:
        return None, None
    s = str(v).strip().rstrip(".")
    partes = [p.strip() for p in s.split("/") if p.strip()]
    uf = None
    # 1) token de 2 letras que seja sigla de UF
    for i, p in enumerate(partes):
        if p.upper() in UFS:
            uf = p.upper()
            partes.pop(i)
            break
    # 2) nome de estado por extenso (ex.: "Bahia/Salvador")
    if uf is None:
        for i, p in enumerate(partes):
            if _sem_acento(p).lower() in ESTADOS:
                uf = ESTADOS[_sem_acento(p).lower()]
                partes.pop(i)
                break
    municipio = "/".join(partes) if partes else None
    if municipio:
        municipio = municipio[:150]
    return municipio, uf


def ler_planilha(caminho, aba):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[aba]
    out = []
    for r in range(LINHA_CABECALHO + 1, ws.max_row + 1):
        instr_raw = ws.cell(row=r, column=_col(COLS["instr"])).value
        if instr_raw in (None, ""):
            continue
        datas = {col: _as_date(ws.cell(row=r, column=_col(letra)).value)
                 for letra, col in MAPA_DATAS.items()}
        municipio, uf = _parse_munuf(ws.cell(row=r, column=_col(COLS["munuf"])).value)
        out.append({
            "linha": r,
            "instr_raw": str(instr_raw).strip(),
            "chave": _digitos(instr_raw),
            "objeto": (ws.cell(row=r, column=_col(COLS["local"])).value or None),
            "municipio": municipio,
            "uf": uf,
            "valor": _parse_valor(ws.cell(row=r, column=_col(COLS["valor"])).value),
            "datas": {c: d for c, d in datas.items() if d is not None},
        })
    return out


def chaves_existentes(conn):
    existentes = set()
    with conn.cursor() as cur:
        cur.execute("SELECT instrumento, tc FROM se_cgpac.aio_solicitacoes")
        for instrumento, tc in cur.fetchall():
            for k in (_digitos(instrumento), _digitos(tc)):
                if k:
                    existentes.add(k)
    return existentes


def importar(arquivo, conn, commit=False, aba=ABA_PADRAO):
    """Insere AIOs ausentes a partir da planilha. Retorna dict de estatísticas.

    `conn` é uma conexão psycopg2 já aberta (não é fechada aqui)."""
    linhas = ler_planilha(arquivo, aba)
    print(f"Planilha: {len(linhas)} linha(s) com instrumento.")

    existentes = chaves_existentes(conn)
    novos, ja_existem, vistos = [], 0, set()
    for ln in linhas:
        k = ln["chave"]
        if k and k in existentes:
            ja_existem += 1
            continue
        if k and k in vistos:        # duplicata dentro da própria planilha
            continue
        if k:
            vistos.add(k)
        novos.append(ln)

    print(f"Já existem no banco (pulados): {ja_existem}")
    print(f"A INSERIR (novos): {len(novos)}\n")
    for ln in novos:
        ds = ", ".join(f"{c}={d.isoformat()}" for c, d in ln["datas"].items()) or "(sem datas)"
        val = f"R$ {ln['valor']:,.2f}" if ln["valor"] is not None else "—"
        print(f"  {ln['instr_raw']:16s} | {str(ln['municipio']):24.24s} {str(ln['uf'] or ''):2s} | {val:>18s} | {ds}")

    stats = {"linhas": len(linhas), "ja_existem": ja_existem, "novos": len(novos), "inseridos": 0}

    if not commit:
        print(f"\n[DRY-RUN] Nada gravado. Rode com --commit para inserir {len(novos)} AIO(s).")
        return stats

    sql = """
            INSERT INTO se_cgpac.aio_solicitacoes
                (instrumento, objeto, municipio, uf, valor_investimento,
                 dt_checklist_cgpac, dt_gabse_oficio, dt_assinatura_sex, email_id)
            VALUES (%(instrumento)s, %(objeto)s, %(municipio)s, %(uf)s, %(valor)s,
                    %(dt_checklist_cgpac)s, %(dt_gabse_oficio)s, %(dt_assinatura_sex)s, %(email_id)s)
            ON CONFLICT (email_id) DO UPDATE SET
                dt_checklist_cgpac = EXCLUDED.dt_checklist_cgpac,
                dt_gabse_oficio    = EXCLUDED.dt_gabse_oficio,
                dt_assinatura_sex  = EXCLUDED.dt_assinatura_sex,
                valor_investimento = EXCLUDED.valor_investimento
        """
    n = 0
    with conn.cursor() as cur:
        for ln in novos:
            params = {
                "instrumento": ln["instr_raw"][:20],
                "objeto": ln["objeto"],
                "municipio": ln["municipio"],
                "uf": ln["uf"],
                "valor": ln["valor"],
                "dt_checklist_cgpac": ln["datas"].get("dt_checklist_cgpac"),
                "dt_gabse_oficio": ln["datas"].get("dt_gabse_oficio"),
                "dt_assinatura_sex": ln["datas"].get("dt_assinatura_sex"),
                "email_id": f"PLANILHA-{ln['chave'] or ln['linha']}",
            }
            cur.execute(sql, params)
            n += 1
    conn.commit()
    stats["inseridos"] = n
    print(f"\n[OK] {n} AIO(s) inserido(s)/atualizado(s) no banco.")
    return stats


def main():
    ap = argparse.ArgumentParser(description="Insere AIOs faltantes a partir da planilha.")
    ap.add_argument("--arquivo", required=True)
    ap.add_argument("--aba", default=ABA_PADRAO)
    ap.add_argument("--commit", action="store_true",
                    help="Grava no banco. Sem a flag, roda em dry-run.")
    args = ap.parse_args()

    conn = get_db_connection()
    try:
        importar(args.arquivo, conn, commit=args.commit, aba=args.aba)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
