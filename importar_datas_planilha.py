#!/usr/bin/env python3
"""
Importa as 3 datas da planilha "Informações AIO.xlsx" (aba ListaAIO) para as
colunas de etapa correspondentes em se_cgpac.aio_solicitacoes.

Mapeamento (confirmado com a CGPAC):
    Planilha col I  "Data de chegada na CGPAC"  -> dt_checklist_cgpac  (etapa 6)
    Planilha col J  "Data de entrega para SE"   -> dt_gabse_oficio     (etapa 7)
    Planilha col K  "Data Assinatura GM"        -> dt_assinatura_sex   (etapa 8)

Casamento por identificador: a coluna D ("Instrumento") da planilha mistura
formatos; normalizamos para dígitos (cortando sufixo "/AAAA") e tentamos casar
contra `instrumento` (6 díg.) e, se não achar, contra `tc` (7 díg.) do banco.

Uso:
    python3 importar_datas_planilha.py --arquivo "/caminho/Informações AIO.xlsx"   # dry-run (não grava)
    python3 importar_datas_planilha.py --arquivo "..." --commit                    # grava no banco
"""

import argparse
import datetime
import pathlib
import re
import sys

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from aio_pipeline import get_db_connection  # noqa: E402

ABA_PADRAO = "ListaAIO"
LINHA_CABECALHO = 2          # cabeçalho na linha 2; dados a partir da 3
COL_INSTRUMENTO = "D"
COL_PROCESSO_SEI = "C"
COL_MUNICIPIO = "E"

# Planilha (letra da coluna) -> coluna de etapa no banco
MAPA_DATAS = {
    "I": "dt_checklist_cgpac",   # Data de chegada na CGPAC   (etapa 6)
    "J": "dt_gabse_oficio",      # Data de entrega para SE    (etapa 7)
    "K": "dt_assinatura_sex",    # Data Assinatura GM         (etapa 8)
}


def _digitos(v):
    """Normaliza um identificador para apenas dígitos, cortando sufixo /AAAA."""
    if v is None:
        return ""
    s = str(v).strip()
    s = s.split("/")[0]          # "985333/2025" -> "985333"
    return re.sub(r"\D", "", s)  # remove pontos, traços, espaços


def _col(letra):
    return openpyxl.utils.column_index_from_string(letra)


def _as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def ler_planilha(caminho, aba):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[aba]
    linhas = []
    for r in range(LINHA_CABECALHO + 1, ws.max_row + 1):
        instr_raw = ws.cell(row=r, column=_col(COL_INSTRUMENTO)).value
        if instr_raw in (None, ""):
            continue
        datas = {col: _as_date(ws.cell(row=r, column=_col(letra)).value)
                 for letra, col in MAPA_DATAS.items()}
        if not any(datas.values()):
            continue
        linhas.append({
            "linha": r,
            "instr_raw": instr_raw,
            "chave": _digitos(instr_raw),
            "processo_sei": ws.cell(row=r, column=_col(COL_PROCESSO_SEI)).value,
            "municipio": ws.cell(row=r, column=_col(COL_MUNICIPIO)).value,
            "datas": datas,
        })
    return linhas


def indexar_banco(conn):
    """Mapa chave-normalizada -> id (por instrumento e por tc)."""
    por_instr, por_tc = {}, {}
    with conn.cursor() as cur:
        cur.execute("SELECT id, instrumento, tc FROM se_cgpac.aio_solicitacoes")
        for _id, instrumento, tc in cur.fetchall():
            ki = _digitos(instrumento)
            kt = _digitos(tc)
            if ki:
                por_instr.setdefault(ki, []).append(_id)
            if kt:
                por_tc.setdefault(kt, []).append(_id)
    return por_instr, por_tc


def casar(linha, por_instr, por_tc):
    """Retorna (id, origem) ou (None, motivo)."""
    chave = linha["chave"]
    if not chave:
        return None, "sem-chave"
    if chave in por_instr:
        ids = por_instr[chave]
        return (ids[0], "instrumento") if len(ids) == 1 else (None, f"ambiguo-instrumento({len(ids)})")
    if chave in por_tc:
        ids = por_tc[chave]
        return (ids[0], "tc") if len(ids) == 1 else (None, f"ambiguo-tc({len(ids)})")
    return None, "nao-encontrado"


def main():
    ap = argparse.ArgumentParser(description="Importa datas da planilha AIO para o banco.")
    ap.add_argument("--arquivo", required=True, help="Caminho do .xlsx da planilha.")
    ap.add_argument("--aba", default=ABA_PADRAO)
    ap.add_argument("--commit", action="store_true",
                    help="Grava no banco. Sem esta flag, roda em dry-run (só relatório).")
    args = ap.parse_args()

    linhas = ler_planilha(args.arquivo, args.aba)
    print(f"Planilha: {len(linhas)} linha(s) com identificador e ao menos 1 data.")

    conn = get_db_connection()
    try:
        por_instr, por_tc = indexar_banco(conn)
        print(f"Banco: {len(por_instr)} instrumento(s) / {len(por_tc)} TC(s) indexados.\n")

        updates = []   # (id, {coluna: date})
        nao_casou = []
        for ln in linhas:
            _id, origem = casar(ln, por_instr, por_tc)
            if _id is None:
                nao_casou.append((ln, origem))
                continue
            datas = {c: d for c, d in ln["datas"].items() if d is not None}
            updates.append((_id, datas, ln, origem))

        print(f"== CASARAM: {len(updates)} | NÃO casaram: {len(nao_casou)} ==\n")
        for _id, datas, ln, origem in updates:
            ds = ", ".join(f"{c}={d.isoformat()}" for c, d in datas.items())
            print(f"  [{origem:11s}] {str(ln['instr_raw']):14s} -> id {_id} | {ds}")

        if nao_casou:
            print("\n-- NÃO casaram (precisam de ajuste manual): --")
            for ln, motivo in nao_casou:
                print(f"  {str(ln['instr_raw']):14s} ({motivo}) | {ln['municipio']}")

        if not args.commit:
            print(f"\n[DRY-RUN] Nada foi gravado. Rode com --commit para aplicar "
                  f"{len(updates)} update(s).")
            return

        n = 0
        with conn.cursor() as cur:
            for _id, datas, _ln, _o in updates:
                if not datas:
                    continue
                sets = ", ".join(f"{c} = %s" for c in datas)
                cur.execute(
                    f"UPDATE se_cgpac.aio_solicitacoes SET {sets} WHERE id = %s",
                    list(datas.values()) + [_id],
                )
                n += cur.rowcount
        conn.commit()
        print(f"\n[OK] {n} linha(s) atualizada(s) no banco.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
